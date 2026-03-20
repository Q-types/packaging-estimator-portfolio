"""
Excel Estimate Processor - Parses legacy PackagePro estimate spreadsheets.

Handles 3 template eras:
  - 2017: Simple box costing (49x6, single sheet, no merged cells)
  - 2020: Multi-panel POB master (99x25, 2 sheets, embedded rate tables)
  - 2024: Customer-facing quotation (32x9, 23 merged cells)

Usage:
    processor = ExcelEstimateProcessor()
    result = processor.process_file("/path/to/estimate.xlsx")
    # result is an EstimateRecord with extracted data
"""

import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class EstimateRecord:
    """Extracted data from a legacy estimate file."""

    # Identification
    estimate_id: Optional[str] = None
    date: Optional[datetime] = None
    company_name: Optional[str] = None
    contact_name: Optional[str] = None
    job_description: Optional[str] = None
    reference: Optional[str] = None

    # Template info
    template_era: Optional[str] = None  # "2017_box", "2020_pob", "2024_quotation"
    file_path: Optional[str] = None
    file_name: Optional[str] = None

    # Financials
    quantities: list[int] = field(default_factory=list)
    prices: list[float] = field(default_factory=list)
    unit_prices: list[float] = field(default_factory=list)
    profit_margin_pct: Optional[float] = None
    total_cost: Optional[float] = None

    # Specifications
    product_type: Optional[str] = None  # "binder", "box", "folder", etc.
    size: Optional[str] = None  # "A4", "A5", etc.
    materials: list[str] = field(default_factory=list)
    operations: list[str] = field(default_factory=list)
    finishing: Optional[str] = None
    printing: Optional[str] = None
    lamination: Optional[str] = None
    packing: Optional[str] = None
    delivery: Optional[str] = None

    # Cost breakdown (from internal costing sheets)
    cost_breakdown: dict[str, float] = field(default_factory=dict)

    # Raw data for ML training
    raw_cells: dict[str, Any] = field(default_factory=dict)

    # Processing metadata
    parse_warnings: list[str] = field(default_factory=list)
    parse_confidence: float = 0.0  # 0-1 confidence in extraction quality

    def to_dict(self) -> dict:
        d = {
            "estimate_id": self.estimate_id,
            "date": self.date.isoformat() if self.date else None,
            "company_name": self.company_name,
            "contact_name": self.contact_name,
            "job_description": self.job_description,
            "reference": self.reference,
            "template_era": self.template_era,
            "file_name": self.file_name,
            "quantities": self.quantities,
            "prices": self.prices,
            "unit_prices": self.unit_prices,
            "profit_margin_pct": self.profit_margin_pct,
            "total_cost": self.total_cost,
            "product_type": self.product_type,
            "size": self.size,
            "materials": self.materials,
            "operations": self.operations,
            "finishing": self.finishing,
            "printing": self.printing,
            "lamination": self.lamination,
            "packing": self.packing,
            "delivery": self.delivery,
            "cost_breakdown": self.cost_breakdown,
            "parse_warnings": self.parse_warnings,
            "parse_confidence": self.parse_confidence,
        }
        return d


def _safe_str(value: Any) -> str:
    """Safely convert a cell value to string."""
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value: Any) -> Optional[float]:
    """Safely convert a cell value to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).replace("£", "").replace(",", "").strip()
        if cleaned:
            return float(cleaned)
    except (ValueError, TypeError):
        pass
    return None


def _safe_int(value: Any) -> Optional[int]:
    """Safely convert a cell value to int."""
    f = _safe_float(value)
    return int(f) if f is not None else None


def _parse_date_from_string(date_str: str) -> Optional[datetime]:
    """Parse date from various formats used in estimates."""
    formats = [
        "%d.%m.%Y",   # 08.10.2024
        "%d/%m/%Y",   # 08/10/2024
        "%d/%m/%y",   # 08/10/24
        "%d.%m.%y",   # 08.10.24
        "%Y-%m-%d",   # 2024-10-08
    ]
    for fmt in formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except (ValueError, AttributeError):
            continue
    return None


def _parse_filename(filename: str) -> dict[str, Optional[str]]:
    """
    Parse estimate filename for metadata.

    Pattern: [ESTIMATE_ID] [DDMMYY] [COMPANY] [DESCRIPTION].xlsx
    Example: "50439 300924 BOXES & PACKAGING 4 PIECE JACKET WITH TRAY.xlsx"
    """
    result: dict[str, Optional[str]] = {
        "estimate_id": None,
        "date": None,
        "description": None,
    }

    name = filename.replace(".xlsx", "").replace(".xls", "")
    parts = name.split(" ", 2)

    if len(parts) >= 1 and parts[0].isdigit():
        result["estimate_id"] = parts[0]

    if len(parts) >= 2:
        # Try to parse DDMMYY date
        date_str = parts[1]
        if len(date_str) == 6 and date_str.isdigit():
            try:
                day = int(date_str[0:2])
                month = int(date_str[2:4])
                year = int(date_str[4:6])
                year += 2000 if year < 100 else 0
                result["date"] = datetime(year, month, day).isoformat()
            except (ValueError, OverflowError):
                pass

    if len(parts) >= 3:
        result["description"] = parts[2]

    return result


def _detect_template_era(ws, num_rows: int, num_cols: int, merged_count: int) -> str:
    """Detect which template era a worksheet belongs to."""
    # 2024 quotation: has merged cells, "ESTIMATE" title, ~32 rows
    cell_a1 = _safe_str(ws.cell(1, 1).value).upper()

    if merged_count > 10 and "ESTIMATE" in cell_a1:
        return "2024_quotation"

    # 2020 POB master: detect by title OR structural pattern
    # Title-based: "PAPER OVER BOARD MASTER" or "MASTER ESTIMATE" in first rows
    for r in range(1, min(5, num_rows + 1)):
        val = _safe_str(ws.cell(r, 1).value).upper()
        if "MASTER" in val or "PAPER OVER BOARD" in val:
            return "2020_pob_master"

    # Structure-based: POB templates have SUB TOTAL A/B/C/D in column 5-6
    # and financial summary (JOB TOTAL, UNIT SELLING PRICE) in cols 5-6 header
    has_subtotals = False
    has_financial_header = False
    for r in range(1, min(10, num_rows + 1)):
        val_e = _safe_str(ws.cell(r, 5).value).upper()
        if "JOB TOTAL" in val_e or "UNIT SELLING" in val_e or "TOTAL COST" in val_e:
            has_financial_header = True
            break
    if has_financial_header and num_rows > 50:
        for r in range(25, min(95, num_rows + 1)):
            val_e = _safe_str(ws.cell(r, 5).value).upper()
            if "SUB TOTAL" in val_e:
                has_subtotals = True
                break
    if has_financial_header and has_subtotals:
        return "2020_pob_master"

    # 2017 box: simple, small, "BOX MAKING QUOTE" or similar
    if num_rows < 60 and num_cols <= 8:
        for r in range(1, min(5, num_rows + 1)):
            val = _safe_str(ws.cell(r, 1).value).upper()
            if "BOX" in val or "QUOTE" in val:
                return "2017_box"

    # Default: try to detect based on size
    if num_cols > 10:
        return "2020_pob_master"
    if merged_count > 5:
        return "2024_quotation"
    return "2017_box"


def _detect_product_type(text: str) -> Optional[str]:
    """Detect product type from description text."""
    text_upper = text.upper()
    type_keywords = {
        "binder": ["BINDER", "POB", "PAPER OVER BOARD"],
        "box": ["BOX", "TRAY", "JACKET", "CASE", "CARTON"],
        "folder": ["FOLDER", "PRESENTATION"],
        "label": ["LABEL", "STICKER"],
        "sign": ["SIGN", "PVC", "FOAM"],
        "divider": ["DIVIDER", "TAB"],
        "packaging": ["PACK", "PACKAGING", "WALLET"],
    }
    for ptype, keywords in type_keywords.items():
        for kw in keywords:
            if kw in text_upper:
                return ptype
    return None


def _detect_size(text: str) -> Optional[str]:
    """Detect standard size from text."""
    text_upper = text.upper()
    sizes = ["A3", "A4", "A5", "A6", "A2", "B5", "B4"]
    for size in sizes:
        if size in text_upper:
            return size
    return None


def _extract_2017_box(ws, record: EstimateRecord) -> None:
    """Extract data from 2017-era box costing template."""
    record.template_era = "2017_box"

    # Header section (rows 1-9) - columns A-B
    for r in range(1, min(10, ws.max_row + 1)):
        label = _safe_str(ws.cell(r, 1).value).upper()
        value = ws.cell(r, 2).value

        if "QUOTE" in label and "NUMBER" in label:
            record.estimate_id = record.estimate_id or _safe_str(value)
        elif "CLIENT" in label or "COMPANY" in label or "CUSTOMER" in label:
            record.company_name = record.company_name or _safe_str(value)
        elif "CONTACT" in label:
            record.contact_name = _safe_str(value)
        elif "DATE" in label:
            if isinstance(value, datetime):
                record.date = value
            elif value:
                record.date = _parse_date_from_string(_safe_str(value)) or record.date
        elif "QUANTIT" in label and "OVERS" not in label:
            qty = _safe_int(value)
            if qty:
                record.quantities.append(qty)
        elif "UNIT" in label and "COST" in label:
            up = _safe_float(value)
            if up:
                record.unit_prices.append(up)
        elif "PROFIT" in label and "MARGIN" in label:
            record.profit_margin_pct = _safe_float(value)

    # Header section - columns E-F (financial summary present in many templates)
    for r in range(1, min(10, ws.max_row + 1)):
        label_e = _safe_str(ws.cell(r, 5).value).upper()
        value_f = ws.cell(r, 6).value

        if "JOB TOTAL" in label_e:
            record.total_cost = record.total_cost or _safe_float(value_f)
        elif "UNIT SELLING" in label_e or ("UNIT" in label_e and "PRICE" in label_e):
            up = _safe_float(value_f)
            if up and up not in record.unit_prices:
                record.unit_prices.append(up)
        elif "TOTAL COST" in label_e:
            record.total_cost = record.total_cost or _safe_float(value_f)
        elif "PROFIT MARGIN" in label_e:
            record.profit_margin_pct = record.profit_margin_pct or _safe_float(value_f)

    # Get unit cost from cols D-E if available
    for r in range(1, min(10, ws.max_row + 1)):
        label_d = _safe_str(ws.cell(r, 4).value).upper()
        if "UNIT" in label_d:
            up = _safe_float(ws.cell(r, 5).value)
            if up and up not in record.unit_prices:
                record.unit_prices.append(up)

    # Cost line items - scan full range of rows, col F (6) for costs
    max_cost_row = min(ws.max_row + 1, 95)
    for r in range(10, max_cost_row):
        label = _safe_str(ws.cell(r, 1).value)
        # Try column F first (most common), fall back to column E
        cost = _safe_float(ws.cell(r, 6).value) or _safe_float(ws.cell(r, 5).value)
        label_upper = label.upper() if label else ""

        # Skip subtotal/summary rows - we want individual items
        if "SUB TOTAL" in label_upper or "JOB LOT" in label_upper:
            continue

        if label and cost and cost > 0:
            record.cost_breakdown[label] = cost
            record.operations.append(label.lower())

    # Footer financial summary (rows 75-95 in cols E-F)
    for r in range(max(1, ws.max_row - 25), ws.max_row + 1):
        label_e = _safe_str(ws.cell(r, 5).value).upper()
        value_f = _safe_float(ws.cell(r, 6).value)

        if "JOB TOTAL" in label_e and value_f:
            record.total_cost = value_f
        elif "TOTAL COST" in label_e and value_f:
            record.total_cost = record.total_cost or value_f
        elif "UNIT SELLING" in label_e and value_f:
            if value_f not in record.unit_prices:
                record.unit_prices.append(value_f)
        elif "PROFIT MARGIN" in label_e and value_f:
            record.profit_margin_pct = value_f

    # Legacy total search (rows 35-50, col A label, col E value)
    if not record.total_cost:
        for r in range(35, min(50, ws.max_row + 1)):
            label = _safe_str(ws.cell(r, 1).value).upper()
            if "TOTAL" in label:
                record.total_cost = _safe_float(ws.cell(r, 5).value)
            elif "MARGIN" in label:
                record.profit_margin_pct = record.profit_margin_pct or _safe_float(ws.cell(r, 2).value)

    record.parse_confidence = 0.7


def _extract_2020_pob(ws, record: EstimateRecord) -> None:
    """Extract data from 2020-era POB master template."""
    record.template_era = "2020_pob_master"

    # Header section (rows 1-8)
    for r in range(1, min(12, ws.max_row + 1)):
        for c in range(1, min(8, ws.max_column + 1)):
            label = _safe_str(ws.cell(r, c).value).upper()
            val_cell = ws.cell(r, c + 1).value if c + 1 <= ws.max_column else None

            if "QUOTE" in label or "ESTIMATE" in label:
                if "NUMBER" in label or "NO" in label:
                    record.estimate_id = record.estimate_id or _safe_str(val_cell)
            elif "DATE" in label:
                if isinstance(val_cell, datetime):
                    record.date = val_cell
                elif val_cell:
                    record.date = _parse_date_from_string(_safe_str(val_cell)) or record.date
            elif "COMPANY" in label or "CLIENT" in label:
                record.company_name = record.company_name or _safe_str(val_cell)
            elif "CONTACT" in label:
                record.contact_name = record.contact_name or _safe_str(val_cell)
            elif "TOTAL" in label and ("COST" in label or "JOB" in label):
                record.total_cost = record.total_cost or _safe_float(val_cell)
            elif "UNIT" in label and ("SELLING" in label or "PRICE" in label):
                up = _safe_float(val_cell)
                if up and up not in record.unit_prices:
                    record.unit_prices.append(up)
            elif ("DESCRIPTION" in label or "JOB" in label) and "TOTAL" not in label:
                record.job_description = record.job_description or _safe_str(val_cell)
            elif "QUANTITY" in label and "OVERS" not in label and "CARTON" not in label:
                qty = _safe_int(val_cell)
                if qty and qty not in record.quantities:
                    record.quantities.append(qty)
            elif "PROFIT" in label and "MARGIN" in label:
                record.profit_margin_pct = _safe_float(val_cell)

    # Extract cost sections (Origination, Finishing, Materials, Carriage)
    section = None
    for r in range(10, min(100, ws.max_row + 1)):
        label = _safe_str(ws.cell(r, 1).value).upper()
        cost_col_e = _safe_float(ws.cell(r, 5).value)
        cost_col_f = _safe_float(ws.cell(r, 6).value)
        cost = cost_col_f or cost_col_e

        if "ORIGINATION" in label or "SUBTOTAL A" in label:
            section = "origination"
        elif "FINISHING" in label or "SUBTOTAL B" in label:
            section = "finishing"
        elif "MATERIAL" in label or "SUBTOTAL C" in label:
            section = "materials"
        elif "CARRIAGE" in label or "SUBTOTAL D" in label:
            section = "carriage"

        if label and cost and cost > 0 and "SUBTOTAL" not in label:
            key = f"{section}:{label}" if section else label
            record.cost_breakdown[key] = cost

        # Detect operations from labels
        op_keywords = {
            "guillotine": "cutting", "cut": "cutting", "trim": "cutting",
            "wrap": "wrapping", "pob machine": "wrapping",
            "crease": "creasing", "drill": "drilling",
            "lamina": "laminating", "foil": "foil_blocking",
            "screen": "screen_printing", "digital": "screen_printing",
            "assembl": "assembly", "rivet": "assembly",
            "liner": "liner_gluing", "glue": "liner_gluing",
        }
        label_lower = label.lower()
        for keyword, op in op_keywords.items():
            if keyword in label_lower and op not in record.operations:
                record.operations.append(op)

    # Extract hourly rates from reference table (columns H-I)
    for r in range(1, min(30, ws.max_row + 1)):
        label_h = _safe_str(ws.cell(r, 8).value)
        rate_i = _safe_float(ws.cell(r, 9).value)
        if label_h and rate_i:
            record.raw_cells[f"rate:{label_h}"] = rate_i

    # Financial summary scan (rows 75+)
    # POB templates put JOB TOTAL / UNIT SELLING PRICE after cost sections.
    # Layout varies: cols E-F (6-col sheets) or cols B-C (9+ col sheets)
    if not record.total_cost:
        for r in range(75, ws.max_row + 1):
            # Check cols E-F
            label_e = _safe_str(ws.cell(r, 5).value).upper()
            value_f = _safe_float(ws.cell(r, 6).value)
            # Check cols B-C
            label_b = _safe_str(ws.cell(r, 2).value).upper()
            value_c = _safe_float(ws.cell(r, 3).value)

            for lbl, val in [(label_e, value_f), (label_b, value_c)]:
                if not val:
                    continue
                if "JOB TOTAL" in lbl:
                    record.total_cost = val
                elif "TOTAL COST" in lbl:
                    record.total_cost = record.total_cost or val
                elif "UNIT SELLING" in lbl:
                    if val not in record.unit_prices:
                        record.unit_prices.append(val)
                elif "PROFIT MARGIN" in lbl:
                    record.profit_margin_pct = record.profit_margin_pct or val

    record.parse_confidence = 0.8


def _extract_2024_quotation(ws, record: EstimateRecord) -> None:
    """Extract data from 2024-era customer-facing quotation."""
    record.template_era = "2024_quotation"

    # Structured rows with labels in col A, values in col B onwards
    for r in range(1, min(35, ws.max_row + 1)):
        label = _safe_str(ws.cell(r, 1).value).upper()
        value = _safe_str(ws.cell(r, 2).value)

        if "ESTIMATE" in label and "NUMBER" in label:
            record.estimate_id = record.estimate_id or value
        elif label == "ESTIMATE NUMBER:" or ("ESTIMATE" in label and not record.estimate_id):
            record.estimate_id = record.estimate_id or value
        elif "GENERATED BY" in label:
            record.contact_name = value  # Internal contact
        elif "DATE" in label:
            record.date = _parse_date_from_string(value) or record.date
        elif "CONTACT" in label:
            record.contact_name = value
        elif "COMPANY" in label:
            record.company_name = record.company_name or value
        elif "REFERENCE" in label:
            record.reference = value
            record.job_description = record.job_description or value
        elif "JOB DESCRIPTION" in label or "DESCRIPTION" in label:
            record.job_description = value
        elif "SIZE" in label:
            record.size = value
        elif "ARTWORK" in label:
            record.raw_cells["artwork"] = value
        elif "PROOF" in label:
            record.raw_cells["proofs"] = value
        elif "PRINTING" in label:
            record.printing = value
        elif "MATERIAL" in label:
            record.materials.append(value)
        elif "LAMINATION" in label or "LAMINA" in label:
            record.lamination = value
        elif "FINISHING" in label:
            record.finishing = value
        elif "PACKING" in label:
            record.packing = value
        elif "DELIVERY" in label or "COLLECTION" in label:
            record.delivery = value

    # Price grid (typically rows 24-27)
    for r in range(20, min(35, ws.max_row + 1)):
        label = _safe_str(ws.cell(r, 1).value).upper()
        if "QUANTITIES" in label or "QUANTITY" in label:
            for c in range(2, min(10, ws.max_column + 1)):
                qty = _safe_int(ws.cell(r, c).value)
                if qty and qty > 0:
                    record.quantities.append(qty)
        elif "PRICE" in label and "UNIT" not in label:
            for c in range(2, min(10, ws.max_column + 1)):
                price = _safe_float(ws.cell(r, c).value)
                if price and price > 0:
                    record.prices.append(price)
        elif "UNIT" in label:
            pass  # "EACH", "PER 1000" etc.

    # Calculate unit prices from prices and quantities
    if record.prices and record.quantities and len(record.prices) == len(record.quantities):
        record.unit_prices = [
            round(p / q, 4) if q > 0 else 0
            for p, q in zip(record.prices, record.quantities)
        ]

    record.parse_confidence = 0.85


class ExcelEstimateProcessor:
    """
    Processes legacy Excel estimate files into structured records.

    Handles template detection, data extraction, and validation.
    """

    def __init__(self):
        self.processed_count = 0
        self.error_count = 0
        self.warnings: list[str] = []

    def process_file(self, file_path: str) -> Optional[EstimateRecord]:
        """
        Process a single Excel estimate file.

        Args:
            file_path: Path to the .xlsx file.

        Returns:
            EstimateRecord with extracted data, or None if processing fails.
        """
        record = EstimateRecord(file_path=file_path, file_name=os.path.basename(file_path))

        # Extract metadata from filename
        file_meta = _parse_filename(os.path.basename(file_path))
        record.estimate_id = file_meta["estimate_id"]
        if file_meta["date"]:
            try:
                record.date = datetime.fromisoformat(file_meta["date"])
            except (ValueError, TypeError):
                pass

        # Detect product type from filename description
        desc = file_meta.get("description", "")
        if desc:
            record.job_description = desc
            record.product_type = _detect_product_type(desc)
            record.size = _detect_size(desc)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        except Exception as e:
            logger.error(f"Failed to open {file_path}: {e}")
            self.error_count += 1
            return None

        try:
            ws = wb.active
            if ws is None:
                wb.close()
                return None

            num_rows = ws.max_row or 0
            num_cols = ws.max_column or 0
            merged_count = len(ws.merged_cells.ranges) if hasattr(ws, 'merged_cells') else 0

            # Detect template era
            era = _detect_template_era(ws, num_rows, num_cols, merged_count)

            # Extract based on detected era
            if era == "2024_quotation":
                _extract_2024_quotation(ws, record)
            elif era == "2020_pob_master":
                _extract_2020_pob(ws, record)
            else:
                _extract_2017_box(ws, record)

            # Process additional sheets
            for sheet_name in wb.sheetnames[1:]:
                extra_ws = wb[sheet_name]
                sheet_label = _safe_str(extra_ws.cell(1, 1).value).upper()
                if "JETRIX" in sheet_label:
                    record.raw_cells["has_jetrix_sheet"] = True

            # Enrich from company name in filename if not found in sheet
            if not record.company_name and desc:
                # Try to extract company name (usually first part before product description)
                record.company_name = desc.split(" - ")[0].strip() if " - " in desc else None

            self.processed_count += 1

        except Exception as e:
            logger.error(f"Error processing {file_path}: {e}")
            record.parse_warnings.append(f"Processing error: {str(e)}")
            record.parse_confidence = 0.1
            self.error_count += 1
        finally:
            wb.close()

        return record

    def process_directory(
        self,
        directory: str,
        limit: Optional[int] = None,
        progress_callback: Optional[callable] = None,
    ) -> list[EstimateRecord]:
        """
        Process all Excel files in a directory.

        Args:
            directory: Path to directory containing .xlsx files.
            limit: Maximum number of files to process (for testing).
            progress_callback: Optional callback(processed, total, current_file).

        Returns:
            List of EstimateRecord objects.
        """
        records = []
        files = sorted([
            f for f in os.listdir(directory)
            if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$')
        ])

        if limit:
            files = files[:limit]

        total = len(files)
        logger.info(f"Processing {total} Excel files from {directory}")

        for i, filename in enumerate(files):
            file_path = os.path.join(directory, filename)

            if progress_callback:
                progress_callback(i + 1, total, filename)

            record = self.process_file(file_path)
            if record:
                records.append(record)

            if (i + 1) % 100 == 0:
                logger.info(f"Processed {i + 1}/{total} files ({self.error_count} errors)")

        logger.info(
            f"Done: {self.processed_count} processed, {self.error_count} errors "
            f"out of {total} files"
        )
        return records

    def get_summary(self, records: list[EstimateRecord]) -> dict:
        """Generate a summary of processed records."""
        product_types: dict[str, int] = {}
        companies: dict[str, int] = {}
        eras: dict[str, int] = {}
        years: dict[int, int] = {}

        for r in records:
            if r.product_type:
                product_types[r.product_type] = product_types.get(r.product_type, 0) + 1
            if r.company_name:
                companies[r.company_name] = companies.get(r.company_name, 0) + 1
            if r.template_era:
                eras[r.template_era] = eras.get(r.template_era, 0) + 1
            if r.date:
                year = r.date.year
                years[year] = years.get(year, 0) + 1

        return {
            "total_records": len(records),
            "processed": self.processed_count,
            "errors": self.error_count,
            "product_types": dict(sorted(product_types.items(), key=lambda x: -x[1])),
            "top_companies": dict(sorted(companies.items(), key=lambda x: -x[1])[:20]),
            "template_eras": eras,
            "estimates_by_year": dict(sorted(years.items())),
            "avg_parse_confidence": (
                sum(r.parse_confidence for r in records) / len(records)
                if records else 0
            ),
        }
